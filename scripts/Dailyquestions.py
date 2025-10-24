import smtplib
import pandas as pd
import random
import os
from email.mime.text import MIMEText
from openpyxl import load_workbook
from datetime import datetime
from dotenv import load_dotenv
import time

# Load environment variables
load_dotenv()

EXCEL_FILE = "LeetCode_Roadmap_Questions.xlsx"
LOG_FILE = "sent_questions_log.txt"

YOUR_EMAIL = os.getenv("EMAIL_USER")
APP_PASSWORD = os.getenv("EMAIL_PASS")
TO_EMAILS = os.getenv("TO_EMAILS").split(",")

# === Load Data ===
df = pd.read_excel(EXCEL_FILE)

if "Solved" not in df.columns:
    df["Solved"] = ""

unsolved = df[df["Solved"] != "Yes"]

if len(unsolved) < 3:
    selected = unsolved
else:
    selected = unsolved.sample(3, random_state=random.randint(1, 10000))

# === Prepare Email ===
msg_content = "📌 Today's 3 Random DSA Questions:\n\n"
for _, row in selected.iterrows():
    msg_content += f"- {row['Question Name']} ({row['Topic']})\n  {row['Link']}\n\n"

msg = MIMEText(msg_content, "plain", "utf-8")
msg["Subject"] = "Today's 3 DSA Questions 🚀"
msg["From"] = YOUR_EMAIL
msg["To"] = ", ".join(TO_EMAILS)

# === Send Email ===
with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
    server.login(YOUR_EMAIL, APP_PASSWORD)
    server.sendmail(YOUR_EMAIL, TO_EMAILS, msg.as_string())

print("✅ Email sent successfully!")

# === Update Excel ===
wb = load_workbook(EXCEL_FILE)
ws = wb.active
for idx in selected.index:
    ws.cell(row=idx+2, column=df.columns.get_loc("Solved")+1).value = "Yes"

# Attempt to save workbook with retries to handle file locks (e.g., Excel open)
def save_workbook_with_retry(workbook, filename, attempts=5, delay_seconds=2):
    for attempt in range(1, attempts + 1):
        try:
            workbook.save(filename)
            return True
        except PermissionError:
            if attempt == attempts:
                base, ext = os.path.splitext(filename)
                alt = f"{base}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
                workbook.save(alt)
                print(f"⚠️ Could not write to {filename}. Saved to {alt} instead. Close the file and rename/merge when convenient.")
                return False
            time.sleep(delay_seconds)

save_workbook_with_retry(wb, EXCEL_FILE)

# === Append to Log File ===
with open(LOG_FILE, "a", encoding="utf-8") as f:
    f.write(f"\n📅 {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
    for _, row in selected.iterrows():
        f.write(f"- {row['Question Name']} ({row['Topic']}) | {row['Link']}\n")
    f.write("\n---\n")

print("✅ Log updated")
